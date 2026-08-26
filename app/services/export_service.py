"""
Export service — generate PDF dan Excel laporan aset.
Digunakan oleh Admin Ruangan dan Admin Divisi.
"""
import io
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def generate_excel(asset_data, room_name, room_code):
    """
    Generate file Excel laporan aset.
    asset_data: list of dict {'asset': Asset, 'last_fmea': FmeaRecord|None}
    Return: BytesIO buffer siap di-send_file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Laporan Aset'

    # Header
    headers = [
        'No', 'Kode Barang', 'Nama Aset', 'Kondisi Aset', 'Hasil FMEA',
        'Kategori FMEA', 'Tanggal Evaluasi', 'Catatan Hasil Tindakan',
        'Rekomendasi'
    ]
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for i, item in enumerate(asset_data, start=1):
        a = item['asset']
        # Context laporan memakai report_fmea=None saat RPN belum tersedia.
        # Fallback menjaga kompatibilitas pemanggil lama yang hanya memberi
        # last_fmea.
        f = item.get('report_fmea', item.get('last_fmea'))
        report_risk_category = item.get('report_risk_category') or (f.risk_category if f else 'rendah')
        ws.append([
            i,
            a.item_code or '-',
            a.asset_name,
            a.condition.replace('_', ' ').title(),
            f.rpn_score if f else '-',
            report_risk_category.title(),
            str(f.evaluation_date) if f else '-',
            item.get('action_note') or '-',
            item.get('recommendation') or '-',
        ])

    # Auto column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_excel_divisi(asset_data, room_stats):
    """
    Generate Excel 2-sheet laporan lintas ruangan.
    asset_data: list of dict {'asset': Asset, 'last_fmea': FmeaRecord|None}
    room_stats: list of dict dengan key room_name, total, baik, perlu, kritis, tidak_layak,
                rpn_rendah, rpn_sedang, rpn_tinggi
    Return: BytesIO buffer
    """
    wb = Workbook()

    # ── Sheet 1: Data Aset ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Data Aset'

    headers = ['No', 'Kode Barang', 'Nama Aset', 'Ruangan', 'Kondisi Aset',
               'Hasil FMEA', 'Kategori FMEA', 'Tgl Evaluasi',
               'Catatan Hasil Tindakan', 'Rekomendasi']
    ws1.append(headers)
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_font = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center')
    for cell in ws1[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center

    fill_map = {
        'baik': PatternFill('solid', fgColor='D9EAD3'),
        'perlu_perhatian': PatternFill('solid', fgColor='FFF2CC'),
        'kritis': PatternFill('solid', fgColor='FCE5CD'),
        'tidak_layak': PatternFill('solid', fgColor='F4CCCC'),
    }

    for i, item in enumerate(asset_data, start=1):
        a = item['asset']
        f = item.get('report_fmea', item.get('last_fmea'))
        report_risk_category = item.get('report_risk_category') or (f.risk_category if f else 'rendah')
        ws1.append([
            i,
            a.item_code or '-',
            a.asset_name,
            a.room.room_name,
            a.condition.replace('_', ' ').title(),
            f.rpn_score if f else '-',
            report_risk_category.title(),
            str(f.evaluation_date) if f else '-',
            item.get('action_note') or '-',
            item.get('ai_recommendation') or '-',
        ])
        row_fill = fill_map.get(a.condition)
        if row_fill:
            for cell in ws1[i + 1]:
                cell.fill = row_fill

    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ── Sheet 2: Statistik per Ruangan ──────────────────────────────────────
    ws2 = wb.create_sheet('Statistik per Ruangan')
    headers2 = ['Ruangan', 'Total Aset', 'Baik', 'Perlu Perhatian', 'Kritis',
                'Tidak Layak', 'RPN Rendah', 'RPN Sedang', 'RPN Tinggi']
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center

    for rs in room_stats:
        ws2.append([
            rs['room_name'], rs['total'],
            rs['baik'], rs['perlu'], rs['kritis'], rs['tidak_layak'],
            rs.get('rpn_rendah', 0), rs.get('rpn_sedang', 0), rs.get('rpn_tinggi', 0),
        ])

    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_pdf(template_string):
    """
    Generate PDF dari HTML string menggunakan xhtml2pdf.
    template_string: hasil render_template HTML
    Return: bytes PDF
    """
    from xhtml2pdf import pisa
    buf = io.BytesIO()
    pisa.CreatePDF(template_string, dest=buf, encoding='utf-8')
    buf.seek(0)
    return buf.read()


def build_filename(prefix, room_code, ext):
    """Helper untuk nama file ekspor."""
    tanggal = datetime.now().strftime('%Y%m%d')
    return f"{prefix}_{room_code}_{tanggal}.{ext}"


def generate_maintenance_excel(logs):
    """Export maintenance history using the client's 17-column raw format."""
    wb = Workbook()
    grouped = defaultdict(list)
    for log in logs:
        grouped[(log.action_date.year if log.action_date else datetime.now().year)].append(log)
    if not grouped:
        grouped[datetime.now().year] = []

    for index, (year, year_logs) in enumerate(sorted(grouped.items())):
        ws = wb.active if index == 0 else wb.create_sheet()
        ws.title = str(year)
        ws.cell(row=1, column=1, value='LAPORAN KERUSAKAN ALAT KESEHATAN MEDIK')
        ws.cell(row=2, column=1, value='RS KHUSUS GIGI DAN MULUT PROVINSI SUMATERA SELATAN')
        ws.cell(row=4, column=1, value='No')
        ws.cell(row=4, column=2, value='Hari / Tanggal')
        ws.cell(row=4, column=3, value='Yang Melaporkan')
        ws.cell(row=4, column=6, value='Kondisi Alat :')
        ws.cell(row=4, column=13, value='Hasil Peninjauan Oleh :')
        ws.cell(row=4, column=17, value='Kesimpulan/Saran')
        for column, value in enumerate([
            'Instalasi/Unit', 'Nama', 'Jabatan', 'Nama Alat', 'Jumlah Alat',
            'Merk/Type', 'Serial Number/SN', 'Kode Barang', 'Lokasi Detail Alat', 'Keluhan',
            'Instalasi/Unit', 'Nama', 'Jabatan', 'Hasil', '',
        ], start=3):
            ws.cell(row=5, column=column, value=value)
        # Keep the same merged multi-row header as the client's raw workbook.
        ws.cell(row=6, column=1, value=None)
        ws.cell(row=6, column=2, value=None)
        ws.cell(row=6, column=17, value=None)
        for no, log in enumerate(sorted(year_logs, key=lambda item: item.action_date or datetime.min.date()), start=1):
            asset = log.asset
            values = [
                no, log.action_date, log.reporter_unit or '—', log.reporter_name or '—',
                log.reporter_position or '—', asset.asset_name,
                f'{asset.quantity} {asset.unit or "unit"}', asset.brand_model,
                asset.serial_number or '—', asset.item_code or '—',
                log.location_detail or (asset.room.room_name if asset.room else '—'),
                log.complaint or '—', log.inspection_unit or '—', log.technician_name or '—',
                log.technician_position or '—', log.result or log.description, log.recommendation or '—',
            ]
            # The source workbook reserves rows 4-6 for the multi-row header.
            data_row = 6 + no
            for column, value in enumerate(values, start=1):
                ws.cell(row=data_row, column=column, value=value)

        _style_raw_maintenance_sheet(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_preventive_excel(records):
    """Export preventive history in the client's checklist columns."""
    wb = Workbook()
    if not records:
        records_by_room = {'Preventive': []}
    else:
        records_by_room = defaultdict(list)
        for record in records:
            room_name = record.room_name_snapshot or (record.asset.room.room_name if record.asset.room else 'Preventive')
            records_by_room[room_name].append(record)

    for index, (room_name, room_records) in enumerate(sorted(records_by_room.items())):
        ws = wb.active if index == 0 else wb.create_sheet()
        ws.title = _safe_sheet_title(room_name)
        ws.merge_cells('A1:I1')
        ws['A1'] = 'PREVENTIVE MAINTENANCE CHECKLIST'
        ws.merge_cells('A2:I2')
        ws['A2'] = 'RUMAH SAKIT KHUSUS GIGI DAN MULUT PROVINSI SUMATERA SELATAN'
        ws['A4'] = f'Ruangan : {room_name}'
        ws['A5'] = 'Tanggal : diekspor dari data SIMASET'
        headers = ['No.', 'Tanggal', 'Nama Alat', 'Merk', 'SN', 'Type', 'Hasil', 'Ket.', 'Kondisi Sistem']
        for column, header in enumerate(headers, start=1):
            ws.cell(row=7, column=column, value=header)
        for no, record in enumerate(sorted(room_records, key=lambda item: item.check_date or datetime.min.date()), start=1):
            asset = record.asset
            ws.append([
                no,
                record.check_date,
                asset.asset_name,
                asset.brand or '—',
                asset.serial_number or '—',
                asset.model or '—',
                record.result,
                record.notes or record.recommendation or '—',
                (record.condition_after or asset.condition or '—').replace('_', ' ').title(),
            ])
        _style_preventive_sheet(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _style_preventive_sheet(ws):
    title_fill = PatternFill('solid', fgColor='0B7896')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    for cell in ws[1] + ws[2]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = title_fill
    ws['A2'].font = Font(italic=True, color='666666')
    for cell in ws[7]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=8):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    widths = [8, 15, 30, 22, 20, 20, 45, 45, 20]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    ws.freeze_panes = 'A8'
    ws.auto_filter.ref = f'A7:I{max(ws.max_row, 7)}'


def _safe_sheet_title(value):
    title = ''.join('-' if char in ':\\/?*[]' else char for char in str(value)).strip() or 'Preventive'
    return title[:31]


def _style_raw_maintenance_sheet(ws):
    group_fill = PatternFill('solid', fgColor='D9EAF7')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    for merge in (
        'A1:Q1', 'A2:Q2', 'A4:A6', 'B4:B6', 'C4:E4', 'C5:C6', 'D5:D6', 'E5:E6',
        'F4:L4', 'F5:F6', 'G5:G6', 'H5:H6', 'I5:I6', 'J5:J6', 'K5:K6', 'L5:L6',
        'M4:P4', 'M5:M6', 'N5:N6', 'O5:O6', 'P5:P6', 'Q4:Q6',
    ):
        ws.merge_cells(merge)
    for cell in ws[1]:
        cell.font = Font(bold=True, size=14)
    for cell in ws[4]:
        cell.fill = group_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=7):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    widths = [8, 18, 18, 25, 20, 28, 15, 22, 20, 20, 28, 42, 18, 25, 20, 55, 45]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    ws.freeze_panes = 'A7'
    ws.auto_filter.ref = f'A5:Q{max(ws.max_row, 5)}'


def generate_kir_pdf(asset, printed_by, base_url):
    """
    Generate PDF Kartu KIR untuk satu aset menggunakan xhtml2pdf.

    Args:
        asset:       instance model Asset
        printed_by:  string nama user yang mencetak
        base_url:    base URL aplikasi (tidak digunakan, QR embed sebagai base64)

    Returns:
        bytes: isi file PDF
    """
    import base64
    from datetime import date
    from flask import render_template
    from xhtml2pdf import pisa
    from app.utils.helpers import generate_qr_code, format_date

    # Generate (atau reuse) QR code
    qr_path = generate_qr_code(asset.id, base_url)

    # xhtml2pdf mendukung data URI base64 — tidak butuh file:// atau server running
    qr_uri = None
    if qr_path:
        try:
            with open(qr_path, 'rb') as f:
                qr_b64 = base64.b64encode(f.read()).decode('utf-8')
            qr_uri = f"data:image/png;base64,{qr_b64}"
        except Exception:
            qr_uri = None

    from app.models.fmea import FmeaRecord
    fmea_terakhir = (FmeaRecord.query
                     .filter_by(asset_id=asset.id)
                     .order_by(FmeaRecord.created_at.desc())
                     .first())

    html_string = render_template(
        'exports/kir_template.html',
        asset=asset,
        fmea_terakhir=fmea_terakhir,
        qr_path=qr_uri,
        printed_by=printed_by,
        printed_at=datetime.now(),
        today_date=date.today(),
        format_date=format_date,
    )

    buf = io.BytesIO()
    pisa.CreatePDF(html_string, dest=buf, encoding='utf-8')
    buf.seek(0)
    return buf.read()
