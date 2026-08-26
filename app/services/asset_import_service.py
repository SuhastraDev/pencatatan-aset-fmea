"""Preview and import identity data from the client's KIB B workbook."""

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
import tempfile
from uuid import UUID, uuid4

from openpyxl import load_workbook
from sqlalchemy import func
from werkzeug.datastructures import FileStorage

from app import db
from app.models.asset import Asset
from app.models.asset_category import AssetCategory
from app.models.room import Room
from app.services.excel_import_service import _asset_text_matches, _build_item_code, _text
from app.utils.helpers import generate_asset_code


IMPORT_DIR = Path(tempfile.gettempdir()) / 'simaset-asset-imports'


@dataclass
class AssetImportRow:
    sheet_name: str
    row_number: int
    item_code: str
    asset_name: str
    specification: str
    brand_type: str
    serial_number: str
    quantity: int | None
    unit: str
    purchase_date: date | None
    purchase_price: Decimal | None
    acquisition_document_number: str
    funding_source: str
    room_name: str
    division_name: str
    asset_condition: str
    asset_status: str
    notes: str
    status: str
    status_label: str
    match_note: str = ''
    matched_asset: Asset | None = None
    target_room_id: int | None = None


@dataclass
class AssetImportPreview:
    rows: list[AssetImportRow] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    ignored_rows: int = 0

    @property
    def total_rows(self):
        return len(self.rows) + self.ignored_rows

    @property
    def matched_count(self):
        return sum(row.status == 'matched' for row in self.rows)

    @property
    def new_asset_count(self):
        return sum(row.status == 'new_asset' for row in self.rows)

    @property
    def importable_count(self):
        return self.matched_count + self.new_asset_count

    @property
    def ambiguous_count(self):
        return sum(row.status == 'ambiguous' for row in self.rows)

    @property
    def unmatched_count(self):
        return sum(row.status == 'unmatched' for row in self.rows)

    @property
    def duplicate_count(self):
        return sum(row.status == 'duplicate' for row in self.rows)

    @property
    def invalid_count(self):
        return sum(row.status == 'invalid' for row in self.rows)

    @property
    def has_blocking_rows(self):
        return self.ambiguous_count > 0 or self.invalid_count > 0


@dataclass
class AssetImportResult:
    assets_created: int = 0
    assets_updated: int = 0
    rows_skipped: int = 0
    ambiguous_rows: int = 0
    unmatched_rows: int = 0


def store_upload(upload: FileStorage):
    IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    token = str(uuid4())
    path = IMPORT_DIR / f'{token}.xlsx'
    upload.save(path)
    return token


def pending_upload_path(token):
    if not token:
        return None
    try:
        UUID(token)
    except (ValueError, TypeError, AttributeError):
        return None

    path = (IMPORT_DIR / f'{token}.xlsx').resolve()
    try:
        path.relative_to(IMPORT_DIR.resolve())
    except ValueError:
        return None
    return path if path.is_file() else None


def remove_upload(token):
    path = pending_upload_path(token)
    if path:
        try:
            path.unlink(missing_ok=True)
        except OSError:
            pass


def build_asset_preview(path, allowed_room_ids=None):
    workbook = load_workbook(path, data_only=True, read_only=False)
    allowed_room_ids = None if allowed_room_ids is None else set(allowed_room_ids)
    preview = AssetImportPreview()
    seen_keys = set()

    sheet_name = 'Lembar1' if 'Lembar1' in workbook.sheetnames else (
        'Table 1' if 'Table 1' in workbook.sheetnames else None
    )
    if not sheet_name:
        raise ValueError('Sheet Lembar1 atau Table 1 tidak ditemukan pada file KIB B.')

    ws = workbook[sheet_name]
    column_map = _build_column_map(ws, header_row=6)
    assets = _scoped_assets(allowed_room_ids)
    rooms = _scoped_rooms(allowed_room_ids)
    for row_number in range(7, ws.max_row + 1):
        row = _parse_row(ws, sheet_name, row_number, column_map)
        if not row:
            preview.ignored_rows += 1
            continue

        row_key = (
            _norm(row.item_code),
            _norm(row.asset_name),
            _norm(row.specification),
            row.quantity,
        )
        if row_key in seen_keys:
            row.status = 'duplicate'
            row.status_label = 'Duplikat file'
            row.match_note = 'Baris identik sudah muncul sebelumnya.'
            preview.rows.append(row)
            continue
        seen_keys.add(row_key)

        matches, match_note = _find_matches(row, assets)
        if len(matches) == 1:
            row.status = 'matched'
            row.status_label = 'Aset cocok'
            row.match_note = match_note
            row.matched_asset = matches[0]
        elif len(matches) > 1:
            row.status = 'ambiguous'
            row.status_label = 'Perlu dipilih'
            row.match_note = 'Lebih dari satu aset cocok; tidak diubah otomatis.'
        else:
            target_room = _resolve_import_room(row, rooms)
            if target_room:
                row.status = 'new_asset'
                row.status_label = 'Aset baru'
                row.target_room_id = target_room.id
                row.match_note = (
                    f'Aset belum ada; akan dibuat di ruangan {target_room.room_name} '
                    'saat konfirmasi.'
                )
            else:
                row.status = 'unmatched'
                row.status_label = 'Belum cocok'
                row.match_note = (
                    'Ruangan pada file tidak cocok dengan ruangan yang dapat diakses '
                    'oleh akun ini; aset tidak dibuat.'
                )
        preview.rows.append(row)

    if preview.unmatched_count:
        preview.warnings.append(
            f'{preview.unmatched_count} baris KIB tidak dapat dibuat karena ruangan pada file '
            'tidak cocok dengan ruangan yang dapat diakses akun ini.'
        )
    if preview.ambiguous_count:
        preview.warnings.append(
            f'{preview.ambiguous_count} baris memiliki lebih dari satu pasangan; perbaiki identitas aset atau import setelah pemetaan.'
        )
    return preview


def commit_asset_import(path, allowed_room_ids=None, created_by=None):
    preview = build_asset_preview(path, allowed_room_ids=allowed_room_ids)
    if preview.has_blocking_rows:
        raise ValueError('Preview KIB masih memiliki baris ambigu atau tidak valid. Periksa dahulu sebelum menyimpan.')

    result = AssetImportResult(
        ambiguous_rows=preview.ambiguous_count,
        unmatched_rows=preview.unmatched_count,
    )
    rooms = {room.id: room for room in _scoped_rooms(allowed_room_ids)}
    category = _get_import_category()
    sequence_cache = {}
    updated_ids = set()
    for row in preview.rows:
        if row.status == 'new_asset' and row.target_room_id in rooms:
            asset = _create_asset_from_row(
                row,
                rooms[row.target_room_id],
                category,
                created_by,
                sequence_cache,
            )
            db.session.add(asset)
            result.assets_created += 1
            continue

        if row.status != 'matched' or not row.matched_asset:
            result.rows_skipped += 1
            continue
        if row.matched_asset.id in updated_ids:
            result.rows_skipped += 1
            continue
        updated_ids.add(row.matched_asset.id)
        if _update_asset(row.matched_asset, row):
            result.assets_updated += 1

    db.session.commit()
    return result


def _scoped_assets(allowed_room_ids):
    query = Asset.query
    if allowed_room_ids is not None:
        if not allowed_room_ids:
            return []
        query = query.filter(Asset.room_id.in_(allowed_room_ids))
    return query.order_by(Asset.id).all()


def _scoped_rooms(allowed_room_ids):
    query = Room.query.filter_by(is_active=True)
    if allowed_room_ids is not None:
        if not allowed_room_ids:
            return []
        query = query.filter(Room.id.in_(allowed_room_ids))
    return query.order_by(Room.id).all()


def _resolve_import_room(row, rooms):
    """Resolve a KIB room only within the current user's allowed rooms."""
    room_key = _norm(row.room_name)
    if not room_key:
        return rooms[0] if len(rooms) == 1 else None

    matches = [
        room for room in rooms
        if room_key in {_norm(room.room_name), _norm(room.room_code)}
    ]
    return matches[0] if len(matches) == 1 else None


def _get_import_category():
    category = AssetCategory.query.filter_by(category_name='Umum').first()
    if category:
        return category
    category = AssetCategory(
        category_name='Umum',
        description='Kategori default untuk aset yang diimpor dari KIB B.',
    )
    db.session.add(category)
    db.session.flush()
    return category


def _create_asset_from_row(row, room, category, created_by, sequence_cache):
    room_code = room.room_code or 'IMP'
    if room_code not in sequence_cache:
        last_code = (
            db.session.query(func.max(Asset.asset_code))
            .filter(Asset.asset_code.like(f'AST-{room_code}-%'))
            .scalar()
        )
        try:
            sequence_cache[room_code] = int(last_code.rsplit('-', 1)[-1]) if last_code else 0
        except (ValueError, AttributeError):
            sequence_cache[room_code] = Asset.query.filter_by(room_id=room.id).count()

    sequence_cache[room_code] += 1
    return Asset(
        asset_code=generate_asset_code(room_code, sequence_cache[room_code]),
        item_code=row.item_code or None,
        asset_name=row.asset_name,
        specification=row.specification or None,
        category=category,
        room=room,
        brand=row.brand_type or '-',
        model='',
        serial_number=row.serial_number or None,
        quantity=row.quantity or 1,
        unit=row.unit or 'unit',
        purchase_date=row.purchase_date,
        purchase_price=row.purchase_price,
        acquisition_document_number=row.acquisition_document_number or None,
        funding_source=row.funding_source or None,
        condition=row.asset_condition,
        status=row.asset_status,
        notes=row.notes or f'Dibuat dari import KIB B sheet {row.sheet_name}, baris {row.row_number}.',
        created_by=created_by,
    )


def _parse_row(ws, sheet_name, row_number, column_map=None):
    column_map = column_map or {}
    asset_name = _text(_cell_value(ws, row_number, column_map, ['Nama Aset'], 10))
    specification = _text(_cell_value(ws, row_number, column_map, ['Spesifikasi', 'Spesifikasi Tambahan'], 21))
    quantity = _to_int(_cell_value(ws, row_number, column_map, ['Jumlah'], 30))
    if asset_name.lower().startswith('contoh'):
        return None
    if not asset_name or not specification or not quantity:
        return None

    return AssetImportRow(
        sheet_name=sheet_name,
        row_number=row_number,
        item_code=(
            _text(_cell_value(ws, row_number, column_map, ['Kode Barang'], None))
            if _norm('Kode Barang') in column_map
            else _build_item_code(ws, row_number)
        ),
        asset_name=asset_name,
        specification=specification,
        brand_type=_text(_cell_value(ws, row_number, column_map, ['Merk/Type', 'Merk / Type', 'Merk', 'Type'], 28)),
        serial_number=_text(_cell_value(ws, row_number, column_map, ['Nomor Seri', 'No Seri'], 29)),
        quantity=quantity,
        unit=_text(_cell_value(ws, row_number, column_map, ['Satuan'], 31)),
        purchase_date=_to_date(_cell_value(ws, row_number, column_map, ['Tanggal Perolehan', 'Tgl Pembelian'], 41)),
        purchase_price=_to_money(_cell_value(ws, row_number, column_map, ['Harga Perolehan', 'Nilai Perolehan', 'Harga Satuan', 'Harga'], 35)) or _to_money(ws.cell(row_number, 33).value),
        acquisition_document_number=_text(_cell_value(ws, row_number, column_map, ['Nomor Dokumen/BAST', 'Dokumen/BAST', 'Dokumen BAST', 'Nomor Dokumen'], 43)),
        funding_source=_text(_cell_value(ws, row_number, column_map, ['Sumber Dana', 'Sumber Dana KIB'], 44)),
        room_name=_text(_cell_value(ws, row_number, column_map, ['Nama Ruangan'], 17)),
        division_name=_text(_cell_value(ws, row_number, column_map, ['Divisi'], 18)),
        asset_condition=_parse_condition(_cell_value(ws, row_number, column_map, ['Kondisi'], 22)),
        asset_status=_parse_status(_cell_value(ws, row_number, column_map, ['Status'], 23)),
        notes=_text(_cell_value(ws, row_number, column_map, ['Catatan', 'Keterangan'], None)),
        status='unmatched',
        status_label='Belum cocok',
    )


def _build_column_map(ws, header_row=6):
    """Map normalized headers to all matching columns for compact and legacy KIB files."""
    columns = {}
    for column in range(1, ws.max_column + 1):
        header = _norm(ws.cell(header_row, column).value)
        if header:
            columns.setdefault(header, []).append(column)
    return columns


def _cell_value(ws, row_number, column_map, aliases, fallback_column):
    """Read the first non-empty aliased header value, then use the legacy column fallback."""
    for alias in aliases:
        for column in column_map.get(_norm(alias), []):
            value = ws.cell(row_number, column).value
            if value not in (None, ''):
                return value
    if fallback_column:
        return ws.cell(row_number, fallback_column).value
    return None


def _find_matches(row, assets):
    serial_number = _norm(row.serial_number)
    if serial_number:
        matches = [a for a in assets if _norm(a.serial_number) == serial_number]
        if matches:
            return matches, 'Cocok berdasarkan nomor seri.'

    item_code = _norm(row.item_code)
    if item_code:
        matches = [a for a in assets if _norm(a.item_code) == item_code]
        if matches:
            return matches, 'Cocok berdasarkan kode barang.'

    name = _norm(row.asset_name)
    specification = _norm(row.specification)
    brand = _norm(row.brand_type)
    exact = [
        a for a in assets
        if _norm(a.asset_name) == name
        and (not specification or _norm(a.specification) == specification)
    ]
    if len(exact) == 1:
        return exact, 'Cocok berdasarkan nama dan spesifikasi.'
    if len(exact) > 1 and brand:
        branded = [a for a in exact if _norm(a.brand_model) == brand]
        if branded:
            return branded, 'Cocok berdasarkan nama, spesifikasi, dan merek/tipe.'
        return exact, 'Nama dan spesifikasi cocok pada beberapa aset.'

    candidates = [a for a in assets if _asset_text_matches(row.asset_name, a.asset_name)]
    if specification:
        candidates = [
            a for a in candidates
            if not _norm(a.specification) or _asset_text_matches(row.specification, a.specification)
        ]
    if len(candidates) > 1 and brand:
        branded = [a for a in candidates if _asset_text_matches(row.brand_type, a.brand_model)]
        if branded:
            candidates = branded
    if len(candidates) == 1:
        return candidates, 'Cocok berdasarkan kemiripan identitas KIB.'
    return candidates, '' if candidates else 'Tidak ditemukan.'


def _update_asset(asset, row):
    values = {
        'item_code': row.item_code or None,
        'asset_name': row.asset_name,
        'specification': row.specification,
        'brand': row.brand_type or None,
        'quantity': row.quantity,
        'unit': row.unit or None,
        'purchase_date': row.purchase_date,
        'purchase_price': row.purchase_price,
        'acquisition_document_number': row.acquisition_document_number or None,
        'funding_source': row.funding_source or None,
        'notes': row.notes or None,
    }
    changed = False
    for key, value in values.items():
        if value in (None, ''):
            continue
        if getattr(asset, key) != value:
            setattr(asset, key, value)
            changed = True
    return changed


def _norm(value):
    return re.sub(r'[^a-z0-9]+', '', _text(value).lower())


def _parse_condition(value):
    key = _norm(value)
    if 'tidaklayak' in key:
        return 'tidak_layak'
    if 'rusakberat' in key or 'kritis' in key:
        return 'kritis'
    if 'perlu' in key or 'rusakringan' in key:
        return 'perlu_perhatian'
    return 'baik'


def _parse_status(value):
    key = _norm(value)
    if 'perbaikan' in key:
        return 'dalam_perbaikan'
    if 'tidakaktif' in key or key in {'nonaktif', 'mati'}:
        return 'tidak_aktif'
    if 'menungguapproval' in key:
        return 'menunggu_approval'
    return 'aktif'


def _to_int(value):
    if isinstance(value, int):
        return value
    match = re.search(r'\d+', _text(value))
    return int(match.group(0)) if match else None


def _to_money(value):
    if value in (None, '', '-'):
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = _text(value).replace('Rp', '').replace('rp', '').replace('.', '').replace(',', '.')
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None
