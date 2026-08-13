from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
import re
import tempfile
from uuid import UUID, uuid4

from openpyxl import load_workbook
from sqlalchemy import func
from werkzeug.datastructures import FileStorage

from app import db
from app.models.asset import Asset
from app.models.asset_category import AssetCategory
from app.models.maintenance_log import MaintenanceLog
from app.models.preventive_maintenance import PreventiveMaintenance
from app.models.room import Room
from app.services.asset_health_service import (
    generate_ai_recommendation,
    infer_condition_from_text,
    recalculate_asset_condition_from_history,
)
from app.services.excel_import_service import (
    _clean_room_name,
    _norm,
    _parse_date,
    _room_name_from_sheet,
    _sheet_date,
    _text,
)
from app.utils.helpers import generate_asset_code


IMPORT_DIR = Path(tempfile.gettempdir()) / 'simaset-preventive-imports'


@dataclass
class PreventiveImportRow:
    sheet_name: str
    row_number: int
    room_name: str
    check_date: date | None
    asset_name: str
    brand: str
    serial_number: str
    model: str
    result: str
    notes: str
    condition_after: str | None
    status: str
    status_label: str
    match_note: str = ''
    matched_asset: Asset | None = None


@dataclass
class PreventiveImportPreview:
    rows: list[PreventiveImportRow] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    ignored_rows: int = 0

    @property
    def total_rows(self):
        return len(self.rows) + self.ignored_rows

    @property
    def matched_count(self):
        return sum(row.status == 'matched' for row in self.rows)

    @property
    def new_asset_count(self):
        return sum(row.status == 'new_asset' for row in self.rows)

    @property
    def duplicate_count(self):
        return sum(row.status == 'duplicate' for row in self.rows)

    @property
    def invalid_count(self):
        return sum(row.status == 'invalid' for row in self.rows)

    @property
    def importable_count(self):
        return self.matched_count + self.new_asset_count

    @property
    def has_blocking_rows(self):
        return self.invalid_count > 0 or any(row.status == 'ambiguous' for row in self.rows)


@dataclass
class PreventiveImportResult:
    assets_created: int = 0
    assets_updated: int = 0
    preventive_created: int = 0
    preventive_skipped: int = 0
    rows_skipped: int = 0


def store_upload(upload: FileStorage):
    """Simpan upload dengan nama acak di folder temporary aplikasi."""
    IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    token = str(uuid4())
    path = IMPORT_DIR / f'{token}.xlsx'
    upload.save(path)
    return token


def pending_upload_path(token):
    """Resolve token session hanya ke file temporary milik aplikasi."""
    if not token:
        return None
    try:
        UUID(token)
    except (ValueError, TypeError, AttributeError):
        return None

    path = (IMPORT_DIR / f'{token}.xlsx').resolve()
    try:
        path.relative_to(IMPORT_DIR.resolve())
    except ValueError:
        return None
    return path if path.is_file() else None


def remove_upload(token):
    path = pending_upload_path(token)
    if path:
        try:
            path.unlink(missing_ok=True)
        except OSError:
            pass


def build_preventive_preview(path, allowed_room_ids=None):
    """Baca workbook tanpa menulis database dan beri status pencocokan setiap baris."""
    workbook = load_workbook(path, data_only=True, read_only=False)
    allowed_room_ids = None if allowed_room_ids is None else set(allowed_room_ids)
    preview = PreventiveImportPreview()
    rooms = _room_lookup(allowed_room_ids)
    seen_rows = set()

    for worksheet in workbook.worksheets:
        raw_room_name = _room_name_from_sheet(worksheet)
        room_name = _clean_room_name(raw_room_name)
        room = rooms.get(_norm(room_name))

        if not room:
            preview.warnings.append(
                f'Sheet "{worksheet.title}" dilewati karena ruangan "{room_name}" '
                'tidak tersedia dalam cakupan akun ini.'
            )
            preview.ignored_rows += max(worksheet.max_row - 8, 0)
            continue

        columns = _preventive_columns(worksheet)
        sheet_date = _sheet_date(worksheet)

        for row_number in range(9, worksheet.max_row + 1):
            values = _read_preventive_row(worksheet, row_number, columns, sheet_date)
            if not values['asset_name'] and not values['result']:
                continue
            if values['asset_name'].strip().upper().startswith('CONTOH'):
                preview.ignored_rows += 1
                continue

            row = PreventiveImportRow(
                sheet_name=worksheet.title,
                row_number=row_number,
                room_name=room.room_name,
                check_date=values['check_date'],
                asset_name=values['asset_name'],
                brand=values['brand'],
                serial_number=values['serial_number'],
                model=values['model'],
                result=values['result'],
                notes=values['notes'],
                condition_after=infer_condition_from_text(values['result'], values['notes']),
                status='new_asset',
                status_label='Aset baru',
            )

            if not row.asset_name or not row.result or not row.check_date:
                row.status = 'invalid'
                row.status_label = 'Perlu diperiksa'
                row.match_note = 'Nama alat, hasil, dan tanggal wajib tersedia.'
                preview.rows.append(row)
                continue

            duplicate_key = _row_key(row)
            if duplicate_key in seen_rows:
                row.status = 'duplicate'
                row.status_label = 'Duplikat file'
                row.match_note = 'Baris dengan aset, tanggal, dan hasil yang sama sudah muncul.'
                preview.rows.append(row)
                continue
            seen_rows.add(duplicate_key)

            match_status, matched_asset, match_note = _match_asset(row, room)
            row.status = match_status
            row.matched_asset = matched_asset
            row.match_note = match_note
            row.status_label = {
                'matched': 'Aset cocok',
                'new_asset': 'Aset baru',
                'ambiguous': 'Perlu diperiksa',
            }.get(match_status, 'Perlu diperiksa')
            preview.rows.append(row)

    return preview


def commit_preventive_import(path, checked_by, allowed_room_ids=None):
    """Simpan hanya baris yang lolos preview dan buat log history yang sama."""
    preview = build_preventive_preview(path, allowed_room_ids=allowed_room_ids)
    if preview.has_blocking_rows:
        raise ValueError(
            'Import belum dapat disimpan karena masih ada baris yang perlu diperiksa.'
        )

    category = _get_import_category()
    room_lookup = _room_lookup(
        None if allowed_room_ids is None else set(allowed_room_ids)
    )
    created_assets = {}
    result = PreventiveImportResult(rows_skipped=preview.duplicate_count)
    sequence_cache = {}

    for row in preview.rows:
        if row.status not in {'matched', 'new_asset'}:
            continue

        room = room_lookup.get(_norm(_clean_room_name(row.room_name)))
        if not room:
            result.rows_skipped += 1
            continue

        asset = row.matched_asset
        if asset is None:
            new_key = _asset_identity_key(row, room.id)
            asset = created_assets.get(new_key)
            if asset is None:
                asset = _create_asset_from_row(
                    row,
                    room,
                    category,
                    checked_by.id,
                    sequence_cache,
                )
                created_assets[new_key] = asset
                result.assets_created += 1
        else:
            if _fill_missing_asset_fields(asset, row):
                result.assets_updated += 1

        db.session.flush()
        duplicate = PreventiveMaintenance.query.filter_by(
            asset_id=asset.id,
            check_date=row.check_date,
            result=row.result,
        ).first()
        if duplicate:
            result.preventive_skipped += 1
            continue

        preventive = PreventiveMaintenance(
            asset_id=asset.id,
            checked_by=checked_by.id,
            check_date=row.check_date,
            room_name_snapshot=room.room_name,
            result=row.result,
            notes=row.notes or None,
            recommendation=row.notes or None,
            condition_after=row.condition_after,
        )
        db.session.add(preventive)

        log = MaintenanceLog(
            asset_id=asset.id,
            logged_by=checked_by.id,
            action_type='preventive_check',
            description=f'Preventive maintenance dari Excel: {row.result}',
            result=row.result,
            recommendation=row.notes or None,
            condition_after=row.condition_after,
            action_date=row.check_date,
        )
        db.session.add(log)
        db.session.flush()

        asset.last_maintenance_date = row.check_date
        recalculate_asset_condition_from_history(asset)
        preventive.ai_recommendation = generate_ai_recommendation(asset, preventive=preventive)
        log.ai_recommendation = preventive.ai_recommendation
        result.preventive_created += 1

    db.session.commit()
    return result


def _preventive_columns(worksheet):
    has_date_column = _text(worksheet.cell(7, 2).value).lower() == 'tanggal'
    return {
        'date': 2 if has_date_column else None,
        'asset_name': 3 if has_date_column else 2,
        'brand': 4 if has_date_column else 3,
        'serial': 5 if has_date_column else 4,
        'model': 6 if has_date_column else 5,
        'result': 7 if has_date_column else 6,
        'notes': 8 if has_date_column else 7,
    }


def _read_preventive_row(worksheet, row_number, columns, sheet_date):
    check_date = sheet_date
    if columns['date']:
        check_date = _parse_date(worksheet.cell(row_number, columns['date']).value) or sheet_date
    return {
        'check_date': check_date,
        'asset_name': _text(worksheet.cell(row_number, columns['asset_name']).value),
        'brand': _text(worksheet.cell(row_number, columns['brand']).value),
        'serial_number': _text(worksheet.cell(row_number, columns['serial']).value),
        'model': _text(worksheet.cell(row_number, columns['model']).value),
        'result': _text(worksheet.cell(row_number, columns['result']).value),
        'notes': _text(worksheet.cell(row_number, columns['notes']).value),
    }


def _room_lookup(allowed_room_ids):
    query = Room.query.filter_by(is_active=True)
    if allowed_room_ids is not None:
        query = query.filter(Room.id.in_(allowed_room_ids))
    rooms = query.all()
    lookup = {}
    for room in rooms:
        lookup[_norm(room.room_name)] = room
        lookup[_norm(_clean_room_name(room.room_name))] = room
    return lookup


def _match_asset(row, room):
    room_assets = Asset.query.filter_by(room_id=room.id).all()
    serial_key = _serial_key(row.serial_number)
    if serial_key:
        serial_matches = [
            asset for asset in room_assets
            if _serial_key(asset.serial_number) == serial_key
        ]
        if len(serial_matches) == 1:
            return 'matched', serial_matches[0], 'Cocok berdasarkan Serial Number.'
        if len(serial_matches) > 1:
            return 'ambiguous', None, 'Serial Number ditemukan pada lebih dari satu aset.'

    name_key = _asset_key(row.asset_name)
    name_matches = [
        asset for asset in room_assets
        if _asset_key(asset.asset_name) == name_key
    ]
    if row.brand:
        brand_matches = [
            asset for asset in name_matches
            if not asset.brand or _asset_key(asset.brand) == _asset_key(row.brand)
        ]
        if brand_matches:
            name_matches = brand_matches
    if row.model:
        model_matches = [
            asset for asset in name_matches
            if not asset.model or _asset_key(asset.model) == _asset_key(row.model)
        ]
        if model_matches:
            name_matches = model_matches

    if len(name_matches) == 1:
        return 'matched', name_matches[0], 'Cocok berdasarkan nama, merk, dan type.'
    if len(name_matches) > 1:
        return 'ambiguous', None, 'Nama alat cocok dengan lebih dari satu aset.'
    return 'new_asset', None, 'Aset belum ada; aset baru akan dibuat saat konfirmasi.'


def _get_import_category():
    category = AssetCategory.query.filter_by(category_name='Umum').first()
    if category:
        return category
    category = AssetCategory(
        category_name='Umum',
        description='Kategori default untuk aset yang diimpor dari checklist preventive.',
    )
    db.session.add(category)
    db.session.flush()
    return category


def _create_asset_from_row(row, room, category, created_by, sequence_cache):
    room_code = room.room_code or 'IMP'
    if room_code not in sequence_cache:
        last_code = (
            db.session.query(func.max(Asset.asset_code))
            .filter(Asset.asset_code.like(f'AST-{room_code}-%'))
            .scalar()
        )
        try:
            sequence_cache[room_code] = int(last_code.rsplit('-', 1)[-1]) if last_code else 0
        except (ValueError, AttributeError):
            sequence_cache[room_code] = Asset.query.filter_by(room_id=room.id).count()

    sequence_cache[room_code] += 1
    asset = Asset(
        asset_code=generate_asset_code(room_code, sequence_cache[room_code]),
        asset_name=row.asset_name,
        category=category,
        room=room,
        brand=row.brand or '-',
        model=row.model or '',
        serial_number=row.serial_number or None,
        quantity=1,
        unit='unit',
        condition='baik',
        status='aktif',
        notes=f'Dibuat dari checklist preventive sheet {row.sheet_name}.',
        created_by=created_by,
    )
    db.session.add(asset)
    return asset


def _fill_missing_asset_fields(asset, row):
    changed = False
    fields = {
        'brand': row.brand,
        'model': row.model,
        'serial_number': row.serial_number,
    }
    for field_name, value in fields.items():
        current_value = _text(getattr(asset, field_name, None))
        if value and current_value.lower() in {'', '-', '—'}:
            setattr(asset, field_name, value)
            changed = True
    return changed


def _row_key(row):
    identity = _serial_key(row.serial_number) or _asset_identity_key(row, row.room_name)
    return identity, row.check_date, _norm(row.result)


def _asset_identity_key(row, room_id):
    return (
        room_id,
        _serial_key(row.serial_number) or _asset_key(row.asset_name),
        _asset_key(row.brand),
        _asset_key(row.model),
    )


def _serial_key(value):
    return re.sub(r'[^a-z0-9]+', '', _text(value).lower())


def _asset_key(value):
    return re.sub(r'[^a-z0-9]+', '', _text(value).lower())
