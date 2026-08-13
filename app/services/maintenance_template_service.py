"""Generate a maintenance history workbook matching the client's layout."""

from datetime import date
from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def generate_maintenance_template(rooms, template_date=None):
    rooms = sorted(list(rooms), key=lambda room: (room.room_name or '').lower())
    template_date = template_date or date.today()
    room_name = rooms[0].room_name if rooms else 'TEMPLATE'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = str(template_date.year)
    _format_sheet(worksheet, room_name, template_date)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _format_sheet(worksheet, room_name, template_date):
    worksheet.merge_cells('A1:Q1')
    worksheet['A1'] = 'LAPORAN KERUSAKAN ALAT KESEHATAN MEDIK'
    worksheet.merge_cells('A2:Q2')
    worksheet['A2'] = 'RS KHUSUS GIGI DAN MULUT PROVINSI SUMATERA SELATAN'

    merges = [
        'A4:A6', 'B4:B6', 'C4:E4', 'C5:C6', 'D5:D6', 'E5:E6',
        'F4:L4', 'F5:F6', 'G5:G6', 'H5:H6', 'I5:I6', 'J5:J6',
        'K5:K6', 'L5:L6', 'M4:P4', 'M5:M6', 'N5:N6', 'O5:O6',
        'P5:P6', 'Q4:Q6',
    ]
    for merge in merges:
        worksheet.merge_cells(merge)

    headers = {
        'A4': 'No',
        'B4': 'Hari / Tanggal',
        'C4': 'Yang Melaporkan',
        'F4': 'Kondisi Alat :',
        'M4': 'Hasil Peninjauan Oleh :',
        'Q4': 'Kesimpulan/Saran',
        'C5': 'Instalasi/Unit',
        'D5': 'Nama',
        'E5': 'Jabatan',
        'F5': 'Nama Alat',
        'G5': 'Jumlah Alat',
        'H5': 'Merk/Type',
        'I5': 'Serial Number/SN',
        'J5': 'Kode Barang',
        'K5': 'Lokasi Detail Alat',
        'L5': 'Keluhan',
        'M5': 'Instalasi/Unit',
        'N5': 'Nama',
        'O5': 'Jabatan',
        'P5': 'Hasil',
    }
    for cell, value in headers.items():
        worksheet[cell] = value

    example = [
        1,
        template_date,
        room_name,
        'CONTOH - Nama Pelapor',
        'Contoh Jabatan',
        'CONTOH - Dental Unit',
        '1 unit',
        'Contoh Merk/Type',
        'CONTOH-SN-001',
        None,
        'Contoh lokasi alat',
        'Contoh keluhan',
        'IPSRS',
        'Contoh Teknisi',
        'Elektromedik',
        'Contoh hasil peninjauan',
        'Hapus baris CONTOH ini, lalu isi data asli.',
    ]
    for column, value in enumerate(example, start=1):
        worksheet.cell(row=7, column=column, value=value)

    _style_sheet(worksheet)


def _style_sheet(worksheet):
    title_fill = PatternFill('solid', fgColor='0B7896')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    example_fill = PatternFill('solid', fgColor='FFF2CC')
    border = Border(bottom=Side(style='thin', color='D9E2F3'))

    for cell in worksheet[1] + worksheet[2]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    worksheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    worksheet['A1'].fill = title_fill
    worksheet['A2'].font = Font(italic=True, color='666666')

    for row in worksheet.iter_rows(min_row=4, max_row=6, min_col=1, max_col=17):
        for cell in row:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

    for cell in worksheet[7]:
        cell.fill = example_fill
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = border
    worksheet['B7'].number_format = 'dd/mm/yy'

    widths = [7, 16, 18, 24, 20, 26, 13, 25, 23, 16, 28, 34, 18, 26, 18, 46, 46]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + column) if column <= 26 else 'A'].width = width
    worksheet.row_dimensions[1].height = 26
    worksheet.row_dimensions[2].height = 22
    worksheet.row_dimensions[4].height = 28
    worksheet.row_dimensions[5].height = 42
    worksheet.row_dimensions[7].height = 68
    worksheet.freeze_panes = 'A7'
    worksheet.auto_filter.ref = 'A6:Q7'
