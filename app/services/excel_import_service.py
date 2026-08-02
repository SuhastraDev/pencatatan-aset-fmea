import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func

from app import db
from app.models.asset import Asset
from app.models.asset_category import AssetCategory
from app.models.division import Division
from app.models.maintenance_log import MaintenanceLog
from app.models.preventive_maintenance import PreventiveMaintenance
from app.models.room import Room
from app.models.user import User
from app.services.asset_health_service import (
    generate_ai_recommendation,
    infer_condition_from_text,
    recalculate_asset_condition_from_history,
)
from app.utils.helpers import generate_asset_code


MONTHS_ID = {
    'januari': 1,
    'februari': 2,
    'maret': 3,
    'april': 4,
    'mei': 5,
    'juni': 6,
    'juli': 7,
    'agustus': 8,
    'september': 9,
    'oktober': 10,
    'november': 11,
    'desember': 12,
}


@dataclass
class ImportStats:
    assets_created: int = 0
    assets_updated: int = 0
    assets_skipped: int = 0
    maintenance_created: int = 0
    maintenance_skipped: int = 0
    preventive_created: int = 0
    preventive_skipped: int = 0
    rooms_created: int = 0
    warnings: list[str] = field(default_factory=list)

    def lines(self):
        return [
            f'assets_created={self.assets_created}',
            f'assets_updated={self.assets_updated}',
            f'assets_skipped={self.assets_skipped}',
            f'maintenance_created={self.maintenance_created}',
            f'maintenance_skipped={self.maintenance_skipped}',
            f'preventive_created={self.preventive_created}',
            f'preventive_skipped={self.preventive_skipped}',
            f'rooms_created={self.rooms_created}',
            f'warnings={len(self.warnings)}',
        ]


class ClientExcelImporter:
    def __init__(self, dry_run=True, default_division_name='Divisi Rawat Jalan', default_room_code=None):
        self.dry_run = dry_run
        self.default_division_name = default_division_name
        self.default_room_code = default_room_code
        self.stats = ImportStats()
        self._room_cache = {}
        self._asset_sequence_cache = {}

    def run(self, history_file=None, kib_file=None, preventive_file=None):
        self._ensure_basics()

        if history_file:
            self.import_history(history_file)
        if preventive_file:
            self.import_preventive(preventive_file)
        if kib_file:
            self.import_kib(kib_file)

        if self.dry_run:
            db.session.rollback()
        else:
            db.session.commit()
        return self.stats

    def import_kib(self, path):
        workbook = load_workbook(_require_file(path), data_only=True, read_only=False)
        fallback_room = self._get_default_room()

        for sheet_name in ['Lembar1', 'Table 1']:
            if sheet_name not in workbook.sheetnames:
                continue
            ws = workbook[sheet_name]
            for row in range(7, ws.max_row + 1):
                item_name = _text(ws.cell(row, 10).value)
                specification = _text(ws.cell(row, 21).value)
                quantity = _int(ws.cell(row, 30).value)
                if not item_name or not specification or not quantity:
                    continue

                item_code = _build_item_code(ws, row)
                asset = self._find_asset(item_code=item_code, asset_name=item_name, room=fallback_room)
                if not asset and not fallback_room:
                    self.stats.assets_skipped += 1
                    continue

                payload = {
                    'item_code': item_code,
                    'asset_name': item_name,
                    'specification': specification,
                    'brand': _text(ws.cell(row, 28).value) or '-',
                    'model': '',
                    'quantity': quantity,
                    'unit': _text(ws.cell(row, 31).value) or 'unit',
                    'purchase_price': _money(ws.cell(row, 35).value) or _money(ws.cell(row, 33).value),
                    'purchase_date': _parse_date(ws.cell(row, 41).value),
                    'acquisition_document_number': _text(ws.cell(row, 43).value),
                    'funding_source': _text(ws.cell(row, 44).value),
                    'notes': f'Diimpor dari KIB sheet {sheet_name}.',
                }
                self._upsert_asset(asset, fallback_room, payload, source='kib')

    def import_history(self, path):
        workbook = load_workbook(_require_file(path), data_only=True, read_only=False)
        for ws in workbook.worksheets:
            for row in range(7, ws.max_row + 1):
                no = ws.cell(row, 1).value
                if not _looks_like_number(no):
                    continue

                action_date = _parse_date(ws.cell(row, 2).value)
                unit = _text(ws.cell(row, 3).value)
                asset_name = _text(ws.cell(row, 6).value)
                if not action_date or not asset_name:
                    self.stats.maintenance_skipped += 1
                    continue

                room = self._find_or_create_room(unit or _text(ws.cell(row, 11).value) or 'IMPORT')
                payload = {
                    'item_code': _text(ws.cell(row, 10).value),
                    'asset_name': asset_name,
                    'brand': _text(ws.cell(row, 8).value) or '-',
                    'model': '',
                    'serial_number': _text(ws.cell(row, 9).value),
                    'quantity': _int(ws.cell(row, 7).value) or 1,
                    'unit': _unit_from_quantity(ws.cell(row, 7).value),
                    'notes': _text(ws.cell(row, 11).value),
                }
                asset = self._find_asset(
                    item_code=payload['item_code'],
                    serial_number=payload['serial_number'],
                    asset_name=asset_name,
                    room=room,
                )
                asset = self._upsert_asset(asset, room, payload, source='history')

                condition_after = infer_condition_from_text(
                    ws.cell(row, 12).value,
                    ws.cell(row, 16).value,
                    ws.cell(row, 17).value,
                )
                log_payload = {
                    'asset': asset,
                    'logged_by': self._import_user().id if not self.dry_run else None,
                    'action_type': 'perbaikan',
                    'description': _join_parts([
                        f'Keluhan: {_text(ws.cell(row, 12).value)}',
                        f'Hasil: {_text(ws.cell(row, 16).value)}',
                        f'Saran: {_text(ws.cell(row, 17).value)}',
                    ]),
                    'reporter_unit': unit,
                    'reporter_name': _text(ws.cell(row, 4).value),
                    'reporter_position': _text(ws.cell(row, 5).value),
                    'complaint': _text(ws.cell(row, 12).value),
                    'inspection_unit': _text(ws.cell(row, 13).value),
                    'technician_name': _text(ws.cell(row, 14).value),
                    'technician_position': _text(ws.cell(row, 15).value),
                    'result': _text(ws.cell(row, 16).value),
                    'recommendation': _text(ws.cell(row, 17).value),
                    'condition_after': condition_after,
                    'action_date': action_date,
                }
                self._create_maintenance_log(log_payload)

    def import_preventive(self, path):
        workbook = load_workbook(_require_file(path), data_only=True, read_only=False)
        for ws in workbook.worksheets:
            room_name = _room_name_from_sheet(ws)
            room = self._find_or_create_room(room_name)
            sheet_date = _sheet_date(ws)
            has_date_column = _text(ws.cell(7, 2).value).lower() == 'tanggal'

            for row in range(9, ws.max_row + 1):
                asset_name_col = 3 if has_date_column else 2
                brand_col = 4 if has_date_column else 3
                serial_col = 5 if has_date_column else 4
                type_col = 6 if has_date_column else 5
                result_col = 7 if has_date_column else 6
                notes_col = 8 if has_date_column else 7

                asset_name = _text(ws.cell(row, asset_name_col).value)
                result = _text(ws.cell(row, result_col).value)
                if not asset_name or not result:
                    continue

                check_date = _parse_date(ws.cell(row, 2).value) if has_date_column else sheet_date
                check_date = check_date or sheet_date or date.today()
                payload = {
                    'asset_name': asset_name,
                    'brand': _text(ws.cell(row, brand_col).value) or '-',
                    'model': _text(ws.cell(row, type_col).value),
                    'serial_number': _text(ws.cell(row, serial_col).value),
                    'quantity': 1,
                    'unit': 'unit',
                    'notes': f'Diimpor dari checklist preventive {ws.title}.',
                }
                asset = self._find_asset(serial_number=payload['serial_number'], asset_name=asset_name, room=room)
                asset = self._upsert_asset(asset, room, payload, source='preventive')

                condition_after = infer_condition_from_text(result, ws.cell(row, notes_col).value)
                preventive_payload = {
                    'asset': asset,
                    'checked_by': self._import_user().id if not self.dry_run else None,
                    'check_date': check_date,
                    'room_name_snapshot': room.room_name,
                    'result': result,
                    'notes': _text(ws.cell(row, notes_col).value),
                    'recommendation': _text(ws.cell(row, notes_col).value),
                    'condition_after': condition_after,
                }
                self._create_preventive(preventive_payload)

    def _ensure_basics(self):
        self._category()
        self._division()
        self._import_user()

    def _category(self):
        category = AssetCategory.query.filter_by(category_name='Umum').first()
        if category:
            return category
        category = AssetCategory(category_name='Umum', description='Kategori default untuk data import.')
        if not self.dry_run:
            db.session.add(category)
            db.session.flush()
        return category

    def _division(self):
        division = Division.query.filter_by(division_name=self.default_division_name).first()
        if division:
            return division
        division = Division(
            division_name=self.default_division_name,
            description='Divisi default untuk data import Excel klien.',
            is_active=True,
        )
        if not self.dry_run:
            db.session.add(division)
            db.session.flush()
        return division

    def _import_user(self):
        user = User.query.filter_by(email='superadmin@rskgm.id').first()
        if user:
            return user
        user = User.query.filter_by(role='super_admin').first() or User.query.first()
        if user:
            return user
        raise RuntimeError('Tidak ada user untuk created_by/logged_by. Jalankan flask seed dulu.')

    def _get_default_room(self):
        if not self.default_room_code:
            return None
        return Room.query.filter_by(room_code=self.default_room_code).first()

    def _find_or_create_room(self, raw_name):
        room_name = _clean_room_name(raw_name)
        key = _norm(room_name)
        if key in self._room_cache:
            return self._room_cache[key]

        existing = Room.query.filter(func.lower(Room.room_name) == room_name.lower()).first()
        if existing:
            self._room_cache[key] = existing
            return existing

        code = _room_code(room_name)
        existing = Room.query.filter_by(room_code=code).first()
        if existing:
            self._room_cache[key] = existing
            return existing

        room = Room(
            room_code=code,
            room_name=room_name,
            floor='-',
            division=self._division(),
            description='Dibuat otomatis dari import Excel klien.',
            is_active=True,
        )
        self.stats.rooms_created += 1
        if not self.dry_run:
            db.session.add(room)
            db.session.flush()
        self._room_cache[key] = room
        return room

    def _find_asset(self, item_code=None, serial_number=None, asset_name=None, room=None):
        serial = _text(serial_number)
        if serial:
            asset = Asset.query.filter(func.lower(Asset.serial_number) == serial.lower()).first()
            if asset:
                return asset

        item_code = _text(item_code)
        if item_code:
            asset = Asset.query.filter(func.lower(Asset.item_code) == item_code.lower()).first()
            if asset:
                return asset

        if asset_name and room and getattr(room, 'id', None):
            return Asset.query.filter(
                Asset.room_id == room.id,
                func.lower(Asset.asset_name) == _text(asset_name).lower(),
            ).first()
        return None

    def _upsert_asset(self, asset, room, payload, source):
        if asset:
            changed = False
            for key, value in payload.items():
                if value in (None, ''):
                    continue
                if not getattr(asset, key, None):
                    setattr(asset, key, value)
                    changed = True
            if changed:
                self.stats.assets_updated += 1
            else:
                self.stats.assets_skipped += 1
            return asset

        asset = Asset(
            asset_code=self._next_asset_code(room),
            item_code=payload.get('item_code'),
            asset_name=payload.get('asset_name') or 'Aset Import',
            specification=payload.get('specification'),
            category=self._category(),
            room=room,
            brand=payload.get('brand') or '-',
            model=payload.get('model') or '',
            serial_number=payload.get('serial_number'),
            quantity=payload.get('quantity') or 1,
            unit=payload.get('unit') or 'unit',
            purchase_date=payload.get('purchase_date'),
            purchase_price=payload.get('purchase_price'),
            acquisition_document_number=payload.get('acquisition_document_number'),
            funding_source=payload.get('funding_source'),
            condition='baik',
            status='aktif',
            notes=payload.get('notes') or f'Dibuat dari import {source}.',
            creator=self._import_user(),
        )
        self.stats.assets_created += 1
        if not self.dry_run:
            db.session.add(asset)
            db.session.flush()
        return asset

    def _next_asset_code(self, room):
        room_code = room.room_code if room and room.room_code else 'IMP'
        if room_code not in self._asset_sequence_cache:
            last_code = (
                db.session.query(func.max(Asset.asset_code))
                .filter(Asset.asset_code.like(f'AST-{room_code}-%'))
                .scalar()
            )
            seq = 0
            if last_code:
                try:
                    seq = int(last_code.rsplit('-', 1)[-1])
                except (ValueError, IndexError):
                    seq = Asset.query.filter_by(room=room).count() if getattr(room, 'id', None) else 0
            self._asset_sequence_cache[room_code] = seq
        self._asset_sequence_cache[room_code] += 1
        return generate_asset_code(room_code, self._asset_sequence_cache[room_code])

    def _create_maintenance_log(self, payload):
        asset = payload['asset']
        if getattr(asset, 'id', None):
            duplicate = MaintenanceLog.query.filter_by(
                asset_id=asset.id,
                action_type=payload['action_type'],
                action_date=payload['action_date'],
                complaint=payload.get('complaint'),
                result=payload.get('result'),
            ).first()
            if duplicate:
                self.stats.maintenance_skipped += 1
                return None

        self.stats.maintenance_created += 1
        if self.dry_run:
            return None

        log = MaintenanceLog(**payload)
        log.ai_recommendation = generate_ai_recommendation(asset, maintenance_log=log)
        db.session.add(log)
        if payload['action_date']:
            asset.last_maintenance_date = payload['action_date']
        recalculate_asset_condition_from_history(asset)
        return log

    def _create_preventive(self, payload):
        asset = payload['asset']
        if getattr(asset, 'id', None):
            duplicate = PreventiveMaintenance.query.filter_by(
                asset_id=asset.id,
                check_date=payload['check_date'],
                result=payload['result'],
            ).first()
            if duplicate:
                self.stats.preventive_skipped += 1
                return None

        self.stats.preventive_created += 1
        if self.dry_run:
            return None

        preventive = PreventiveMaintenance(**payload)
        preventive.ai_recommendation = generate_ai_recommendation(asset, preventive=preventive)
        db.session.add(preventive)
        db.session.add(MaintenanceLog(
            asset=asset,
            logged_by=payload['checked_by'],
            action_type='preventive_check',
            description=f'Preventive maintenance: {payload["result"]}',
            result=payload['result'],
            recommendation=payload.get('recommendation'),
            condition_after=payload.get('condition_after'),
            ai_recommendation=preventive.ai_recommendation,
            action_date=payload['check_date'],
        ))
        asset.last_maintenance_date = payload['check_date']
        recalculate_asset_condition_from_history(asset)
        return preventive


def _require_file(path):
    if not path:
        raise ValueError('Path file kosong.')
    resolved = Path(path)
    if not resolved.exists():
        raise FileNotFoundError(f'File tidak ditemukan: {resolved}')
    return resolved


def _text(value):
    if value is None:
        return ''
    text = str(value).strip()
    return '' if text in {'-', '—', 'None'} else re.sub(r'\s+', ' ', text)


def _norm(value):
    return re.sub(r'[^a-z0-9]+', '', _text(value).lower())


def _looks_like_number(value):
    if isinstance(value, (int, float)):
        return True
    return _text(value).isdigit()


def _int(value):
    if isinstance(value, int):
        return value
    text = _text(value).lower()
    match = re.search(r'\d+', text)
    return int(match.group(0)) if match else None


def _unit_from_quantity(value):
    text = _text(value)
    match = re.search(r'[a-zA-Z]+', text)
    return match.group(0) if match else 'unit'


def _money(value):
    if value is None or value == '-':
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = _text(value)
    text = text.replace('Rp', '').replace('rp', '').strip()
    text = text.replace('.', '').replace(',', '.')
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value).lower()
    if not text:
        return None
    text = re.sub(r'^[a-z]+,\s*', '', text)
    text = text.replace('tanggal', '').replace(':', ' ').strip()

    for fmt in ('%d/%m/%y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    match = re.search(r'(\d{1,2})\s+([a-z]+)\s+(\d{4})', text)
    if match:
        day, month_name, year = match.groups()
        month = MONTHS_ID.get(month_name)
        if month:
            return date(int(year), month, int(day))
    return None


def _build_item_code(ws, row):
    parts = []
    for col in range(1, 9):
        value = ws.cell(row, col).value
        if value in (None, ''):
            continue
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        parts.append(str(value).strip())
    return '.'.join(parts)


def _room_name_from_sheet(ws):
    for row in range(1, min(ws.max_row, 8) + 1):
        value = _text(ws.cell(row, 1).value)
        if value.lower().startswith('ruangan'):
            return value.split(':', 1)[-1].strip()
    return ws.title


def _sheet_date(ws):
    for row in range(1, min(ws.max_row, 8) + 1):
        value = _text(ws.cell(row, 1).value)
        if value.lower().startswith('tanggal'):
            return _parse_date(value)
    return None


def _clean_room_name(raw_name):
    text = _text(raw_name)
    if not text:
        return 'Import'
    aliases = {
        '160': 'IGD',
        'igd': 'IGD',
        'icu': 'ICU',
        'bedah': 'BEDAH',
        'poli vip': 'POLI VIP',
        'poli umum': 'Poli Umum',
        'rawat jalan': 'Rawat Jalan',
        'rawat jalan poli umum': 'Poli Umum',
        'radiologi': 'Radiologi',
    }
    if text.lower() in aliases:
        return aliases[text.lower()]
    if text.isdigit():
        return f'Ruangan {text}'
    return aliases.get(text.lower(), text)


def _room_code(room_name):
    clean = re.sub(r'[^A-Z0-9]+', '-', room_name.upper()).strip('-')
    clean = clean[:12] or 'IMPORT'
    base = f'IMP-{clean}'[:20]
    code = base
    suffix = 1
    while Room.query.filter_by(room_code=code).first():
        suffix += 1
        code = f'{base[:17]}{suffix:02d}'
    return code


def _join_parts(parts):
    return ' | '.join(part for part in parts if part and not part.endswith(': '))
