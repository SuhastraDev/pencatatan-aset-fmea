"""Preview and import the client's maintenance history workbook."""

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
import re
import tempfile
from uuid import UUID, uuid4

from openpyxl import load_workbook
from sqlalchemy import func
from werkzeug.datastructures import FileStorage

from app import db
from app.models.asset import Asset
from app.models.asset_category import AssetCategory
from app.models.maintenance_log import MaintenanceLog
from app.models.room import Room
from app.services.asset_health_service import (
    generate_ai_recommendation,
    infer_condition_from_text,
    recalculate_asset_condition_from_history,
)
from app.services.excel_import_service import (
    _clean_room_name,
    _norm,
    _parse_date,
    _text,
)
from app.utils.helpers import generate_asset_code


IMPORT_DIR = Path(tempfile.gettempdir()) / 'simaset-maintenance-imports'


@dataclass
class MaintenanceImportRow:
    sheet_name: str
    row_number: int
    room_name: str
    raw_unit: str
    action_date: date | None
    reporter_name: str
    reporter_position: str
    asset_name: str
    quantity: int
    brand_type: str
    serial_number: str
    item_code: str
    location_detail: str
    complaint: str
    inspection_unit: str
    technician_name: str
    technician_position: str
    result: str
    recommendation: str
    condition_after: str | None
    status: str
    status_label: str
    match_note: str = ''
    matched_asset: Asset | None = None


@dataclass
class MaintenanceImportPreview:
    rows: list[MaintenanceImportRow] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    ignored_rows: int = 0

    @property
    def total_rows(self):
        return len(self.rows) + self.ignored_rows

    @property
    def matched_count(self):
        return sum(row.status == 'matched' for row in self.rows)

    @property
    def new_asset_count(self):
        return sum(row.status == 'new_asset' for row in self.rows)

    @property
    def duplicate_count(self):
        return sum(row.status == 'duplicate' for row in self.rows)

    @property
    def invalid_count(self):
        return sum(row.status == 'invalid' for row in self.rows)

    @property
    def importable_count(self):
        return self.matched_count + self.new_asset_count

    @property
    def has_blocking_rows(self):
        return self.invalid_count > 0 or any(
            row.status == 'ambiguous' for row in self.rows
        )


@dataclass
class MaintenanceImportResult:
    assets_created: int = 0
    assets_updated: int = 0
    logs_created: int = 0
    logs_skipped: int = 0
    rows_skipped: int = 0


def store_upload(upload: FileStorage):
    """Save an upload under a random temporary token."""
    IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    token = str(uuid4())
    path = IMPORT_DIR / f'{token}.xlsx'
    upload.save(path)
    return token


def pending_upload_path(token):
    """Resolve a session token only to a file in the import temp directory."""
    if not token:
        return None
    try:
        UUID(token)
    except (ValueError, TypeError, AttributeError):
        return None

    path = (IMPORT_DIR / f'{token}.xlsx').resolve()
    try:
        path.relative_to(IMPORT_DIR.resolve())
    except ValueError:
        return None
    return path if path.is_file() else None


def remove_upload(token):
    path = pending_upload_path(token)
    if path:
        try:
            path.unlink(missing_ok=True)
        except OSError:
            pass


def build_maintenance_preview(path, allowed_room_ids=None):
    """Read the workbook without writing to the database."""
    workbook = load_workbook(path, data_only=True, read_only=False)
    allowed_room_ids = None if allowed_room_ids is None else set(allowed_room_ids)
    preview = MaintenanceImportPreview()
    rooms = _room_lookup(allowed_room_ids)
    seen_log_keys = set()
    skipped_units = set()

    for worksheet in workbook.worksheets:
        for row_number in range(7, worksheet.max_row + 1):
            values = _read_history_row(worksheet, row_number)
            if not _has_history_values(values):
                continue
            if _is_untouched_template_row(values):
                preview.ignored_rows += 1
                preview.warnings.append(
                    f'Sheet "{worksheet.title}" baris {row_number} adalah baris contoh dan dilewati. '
                    'Ganti seluruh isinya dengan data maintenance sebenarnya.'
                )
                continue

            raw_unit = values['reporter_unit']
            room_name = _clean_room_name(raw_unit) if raw_unit else ''
            room = rooms.get(_norm(room_name))
            if not room:
                preview.ignored_rows += 1
                skipped_units.add(raw_unit or '(kosong)')
                continue

            row = _build_preview_row(
                worksheet.title,
                row_number,
                values,
                room.room_name,
            )

            if row.asset_name.strip().upper().startswith('CONTOH'):
                row.status = 'invalid'
                row.status_label = 'Perlu diperiksa'
                row.match_note = 'Hapus kata CONTOH pada Nama Alat dan isi nama alat sebenarnya.'
                preview.rows.append(row)
                continue

            if not row.action_date or not row.asset_name:
                row.status = 'invalid'
                row.status_label = 'Perlu diperiksa'
                row.match_note = 'Tanggal dan nama alat wajib tersedia.'
                preview.rows.append(row)
                continue

            duplicate_key = _log_key(row)
            if duplicate_key in seen_log_keys:
                row.status = 'duplicate'
                row.status_label = 'Duplikat file'
                row.match_note = 'Riwayat dengan aset, tanggal, dan hasil yang sama sudah muncul.'
                preview.rows.append(row)
                continue
            seen_log_keys.add(duplicate_key)

            match_status, matched_asset, match_note = _match_asset(row, room)
            row.status = match_status
            row.matched_asset = matched_asset
            row.match_note = match_note
            row.status_label = {
                'matched': 'Aset cocok',
                'new_asset': 'Aset baru',
                'ambiguous': 'Perlu diperiksa',
            }.get(match_status, 'Perlu diperiksa')
            preview.rows.append(row)

    if skipped_units:
        units = ', '.join(sorted(skipped_units))
        preview.warnings.append(
            f'{preview.ignored_rows} baris dilewati karena unit/ruangan tidak ada '
            f'dalam cakupan akun atau belum terdaftar: {units}.'
        )
    return preview


def commit_maintenance_import(path, logged_by, allowed_room_ids=None):
    """Persist only rows that pass the preview validation."""
    preview = build_maintenance_preview(path, allowed_room_ids=allowed_room_ids)
    if preview.has_blocking_rows:
        raise ValueError(
            'Import belum dapat disimpan karena masih ada baris yang perlu diperiksa.'
        )

    category = _get_import_category()
    rooms = _room_lookup(
        None if allowed_room_ids is None else set(allowed_room_ids)
    )
    created_assets = {}
    sequence_cache = {}
    result = MaintenanceImportResult(rows_skipped=preview.duplicate_count)

    for row in preview.rows:
        if row.status not in {'matched', 'new_asset'}:
            continue

        room = rooms.get(_norm(_clean_room_name(row.room_name)))
        if not room:
            result.rows_skipped += 1
            continue

        asset = row.matched_asset
        if asset is None:
            asset_key = _asset_identity_key(row, room.id)
            asset = created_assets.get(asset_key)
            if asset is None:
                asset = _create_asset_from_row(
                    row,
                    room,
                    category,
                    logged_by.id,
                    sequence_cache,
                )
                created_assets[asset_key] = asset
                result.assets_created += 1
        elif _fill_missing_asset_fields(asset, row):
            result.assets_updated += 1

        db.session.flush()
        duplicate = MaintenanceLog.query.filter_by(
            asset_id=asset.id,
            action_type='perbaikan',
            action_date=row.action_date,
            complaint=row.complaint or None,
            result=row.result or None,
        ).first()
        if duplicate:
            result.logs_skipped += 1
            continue

        log = MaintenanceLog(
            asset_id=asset.id,
            logged_by=logged_by.id,
            action_type='perbaikan',
            description=_build_description(row),
            reporter_unit=row.raw_unit or None,
            reporter_name=row.reporter_name or None,
            reporter_position=row.reporter_position or None,
            complaint=row.complaint or None,
            inspection_unit=row.inspection_unit or None,
            technician_name=row.technician_name or None,
            technician_position=row.technician_position or None,
            result=row.result or None,
            recommendation=row.recommendation or None,
            condition_after=row.condition_after,
            action_date=row.action_date,
        )
        db.session.add(log)
        db.session.flush()

        if not asset.last_maintenance_date or row.action_date > asset.last_maintenance_date:
            asset.last_maintenance_date = row.action_date
        recalculate_asset_condition_from_history(asset)
        log.ai_recommendation = generate_ai_recommendation(
            asset,
            maintenance_log=log,
        )
        result.logs_created += 1

    db.session.commit()
    return result


def _read_history_row(worksheet, row_number):
    values = [worksheet.cell(row_number, column).value for column in range(1, 18)]
    return {
        'number': values[0],
        'action_date': _parse_date(values[1]),
        'reporter_unit': _text(values[2]),
        'reporter_name': _text(values[3]),
        'reporter_position': _text(values[4]),
        'asset_name': _text(values[5]),
        'quantity': _parse_quantity(values[6]),
        'brand_type': _text(values[7]),
        'serial_number': _text(values[8]),
        'item_code': _text(values[9]),
        'location_detail': _text(values[10]),
        'complaint': _text(values[11]),
        'inspection_unit': _text(values[12]),
        'technician_name': _text(values[13]),
        'technician_position': _text(values[14]),
        'result': _text(values[15]),
        'recommendation': _text(values[16]),
    }


def _build_preview_row(sheet_name, row_number, values, room_name):
    condition = infer_condition_from_text(
        values['complaint'],
        values['result'],
        values['recommendation'],
    )
    return MaintenanceImportRow(
        sheet_name=sheet_name,
        row_number=row_number,
        room_name=room_name,
        raw_unit=values['reporter_unit'],
        action_date=values['action_date'],
        reporter_name=values['reporter_name'],
        reporter_position=values['reporter_position'],
        asset_name=values['asset_name'],
        quantity=values['quantity'],
        brand_type=values['brand_type'],
        serial_number=values['serial_number'],
        item_code=values['item_code'],
        location_detail=values['location_detail'],
        complaint=values['complaint'],
        inspection_unit=values['inspection_unit'],
        technician_name=values['technician_name'],
        technician_position=values['technician_position'],
        result=values['result'],
        recommendation=values['recommendation'],
        condition_after=condition,
        status='new_asset',
        status_label='Aset baru',
    )


def _has_history_values(values):
    return any(
        values[key]
        for key in (
            'action_date',
            'reporter_unit',
            'reporter_name',
            'asset_name',
            'complaint',
            'result',
            'recommendation',
        )
    )


def _is_untouched_template_row(values):
    return (
        _norm(values['asset_name']) == _norm('CONTOH - Dental Unit')
        and _norm(values['reporter_name']) == _norm('CONTOH - Nama Pelapor')
        and _norm(values['reporter_position']) == _norm('Contoh Jabatan')
        and _norm(values['brand_type']) == _norm('Contoh Merk/Type')
        and _norm(values['serial_number']) == _norm('CONTOH-SN-001')
        and _norm(values['location_detail']) == _norm('Contoh lokasi alat')
        and _norm(values['complaint']) == _norm('Contoh keluhan')
        and _norm(values['inspection_unit']) == _norm('IPSRS')
        and _norm(values['technician_name']) == _norm('Contoh Teknisi')
        and _norm(values['technician_position']) == _norm('Elektromedik')
        and _norm(values['result']) == _norm('Contoh hasil peninjauan')
        and _norm(values['recommendation']) == _norm(
            'Hapus baris CONTOH ini, lalu isi data asli.'
        )
    )


def _room_lookup(allowed_room_ids):
    query = Room.query.filter_by(is_active=True)
    if allowed_room_ids is not None:
        query = query.filter(Room.id.in_(allowed_room_ids))
    lookup = {}
    for room in query.all():
        lookup[_norm(room.room_name)] = room
        lookup[_norm(_clean_room_name(room.room_name))] = room
    return lookup


def _match_asset(row, room):
    assets = Asset.query.filter_by(room_id=room.id).all()

    serial_key = _serial_key(row.serial_number)
    if serial_key:
        matches = [asset for asset in assets if _serial_key(asset.serial_number) == serial_key]
        if len(matches) == 1:
            return 'matched', matches[0], 'Cocok berdasarkan Serial Number.'
        if len(matches) > 1:
            return 'ambiguous', None, 'Serial Number ditemukan pada lebih dari satu aset.'

    item_code_key = _norm(row.item_code)
    if item_code_key:
        matches = [asset for asset in assets if _norm(asset.item_code) == item_code_key]
        if len(matches) == 1:
            return 'matched', matches[0], 'Cocok berdasarkan Kode Barang.'
        if len(matches) > 1:
            return 'ambiguous', None, 'Kode Barang ditemukan pada lebih dari satu aset.'

    name_key = _asset_key(row.asset_name)
    matches = [asset for asset in assets if _asset_key(asset.asset_name) == name_key]
    if row.brand_type:
        branded = [asset for asset in matches if _brand_type_matches(asset, row.brand_type)]
        if branded:
            matches = branded

    if len(matches) == 1:
        return 'matched', matches[0], 'Cocok berdasarkan Nama Alat dan Merk/Type.'
    if len(matches) > 1:
        return 'ambiguous', None, 'Nama Alat cocok dengan lebih dari satu aset.'
    return 'new_asset', None, 'Aset belum ada; aset baru akan dibuat saat konfirmasi.'


def _brand_type_matches(asset, brand_type):
    incoming = _asset_key(brand_type)
    existing = _asset_key(asset.brand_model)
    if not incoming or not existing:
        return True
    return incoming == existing or incoming in existing or existing in incoming


def _get_import_category():
    category = AssetCategory.query.filter_by(category_name='Umum').first()
    if category:
        return category
    category = AssetCategory(
        category_name='Umum',
        description='Kategori default untuk data yang diimpor dari riwayat maintenance.',
    )
    db.session.add(category)
    db.session.flush()
    return category


def _create_asset_from_row(row, room, category, created_by, sequence_cache):
    room_code = room.room_code or 'IMP'
    if room_code not in sequence_cache:
        last_code = (
            db.session.query(func.max(Asset.asset_code))
            .filter(Asset.asset_code.like(f'AST-{room_code}-%'))
            .scalar()
        )
        try:
            sequence_cache[room_code] = int(last_code.rsplit('-', 1)[-1]) if last_code else 0
        except (ValueError, AttributeError):
            sequence_cache[room_code] = Asset.query.filter_by(room_id=room.id).count()

    sequence_cache[room_code] += 1
    location_note = f' Lokasi detail: {row.location_detail}.' if row.location_detail else ''
    return Asset(
        asset_code=generate_asset_code(room_code, sequence_cache[room_code]),
        item_code=row.item_code or None,
        asset_name=row.asset_name,
        category=category,
        room=room,
        brand=row.brand_type or '-',
        model='',
        serial_number=row.serial_number or None,
        quantity=row.quantity or 1,
        unit='unit',
        condition='baik',
        status='aktif',
        notes=f'Dibuat dari riwayat maintenance sheet {row.sheet_name}.{location_note}',
        created_by=created_by,
    )


def _fill_missing_asset_fields(asset, row):
    changed = False
    values = {
        'item_code': row.item_code,
        'brand': row.brand_type,
        'serial_number': row.serial_number,
    }
    for field_name, value in values.items():
        current = _text(getattr(asset, field_name, None))
        if value and current.lower() in {'', '-', '—'}:
            setattr(asset, field_name, value)
            changed = True
    if row.location_detail and not _text(asset.notes):
        asset.notes = f'Lokasi detail: {row.location_detail}'
        changed = True
    return changed


def _asset_identity_key(row, room_id):
    return (
        room_id,
        _serial_key(row.serial_number) or _asset_key(row.asset_name),
        _asset_key(row.brand_type),
    )


def _log_key(row):
    return (
        _asset_identity_key(row, row.room_name),
        row.action_date,
        _norm(row.complaint),
        _norm(row.result),
    )


def _build_description(row):
    parts = []
    if row.location_detail:
        parts.append(f'Lokasi detail: {row.location_detail}')
    if row.complaint:
        parts.append(f'Keluhan: {row.complaint}')
    if row.result:
        parts.append(f'Hasil: {row.result}')
    if row.recommendation:
        parts.append(f'Saran: {row.recommendation}')
    return ' | '.join(parts) or f'Riwayat maintenance dari Excel sheet {row.sheet_name}.'


def _parse_quantity(value):
    match = re.search(r'\d+', str(value or ''))
    return int(match.group(0)) if match else 1


def _serial_key(value):
    return re.sub(r'[^a-z0-9]+', '', _text(value).lower())


def _asset_key(value):
    return re.sub(r'[^a-z0-9]+', '', _text(value).lower())
