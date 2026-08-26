"""Generate a compact KIB B template compatible with the asset import parser."""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


KIB_HEADERS = [
    'Kode Barang', 'Nama Aset', 'Nama Ruangan', 'Divisi', 'Merk/Type',
    'Nomor Seri', 'Jumlah', 'Satuan', 'Spesifikasi', 'Kondisi', 'Status',
    'Tanggal Perolehan', 'Harga Perolehan', 'Nomor Dokumen/BAST',
    'Sumber Dana', 'Catatan',
]


def generate_asset_template(template_date=None):
    """Return a compact, web-aligned KIB B workbook whose data starts at row 7."""
    template_date = template_date or date.today()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Lembar1'

    last_column = _column_name(len(KIB_HEADERS))
    worksheet.merge_cells(f'A1:{last_column}1')
    worksheet['A1'] = 'TEMPLATE IMPORT IDENTITAS ASET KIB B'
    worksheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    worksheet['A1'].fill = PatternFill('solid', fgColor='0B7896')
    worksheet['A1'].alignment = Alignment(horizontal='center')

    worksheet.merge_cells(f'A2:{last_column}2')
    worksheet['A2'] = 'Rumah Sakit Khusus Gigi dan Mulut Provinsi Sumatera Selatan'
    worksheet['A2'].font = Font(italic=True, color='666666')
    worksheet['A2'].alignment = Alignment(horizontal='center')

    worksheet.merge_cells(f'A4:{last_column}4')
    worksheet['A4'] = (
        'Isi atau salin data mulai baris 7. Baris CONTOH boleh diganti atau dibiarkan karena akan diabaikan otomatis. '
        'Kolom bertanda (*) wajib: Nama Aset, Nama Ruangan, Jumlah, dan Spesifikasi.'
    )
    worksheet['A4'].alignment = Alignment(wrap_text=True)
    worksheet['A4'].font = Font(bold=True, color='404040')

    required_headers = {'Nama Aset', 'Nama Ruangan', 'Jumlah', 'Spesifikasi'}
    for column, header in enumerate(KIB_HEADERS, start=1):
        cell = worksheet.cell(row=6, column=column, value=header)
        if header in required_headers:
            cell.value = f'{header} (*)'
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E79')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    sample = [
        '1.3.2.7.1.1104.1', 'CONTOH - Patient Monitor', 'CONTOH RUANGAN',
        'CONTOH DIVISI', 'Contoh Merk / Type', 'CONTOH-SN-001', 1, 'unit',
        'Alat monitor pasien dengan layar digital', 'Baik', 'Aktif', template_date,
        25000000, 'CONTOH-BAST-001', 'APBD',
        'Ganti isi baris ini dengan data aset Anda.',
    ]
    for column, value in enumerate(sample, start=1):
        cell = worksheet.cell(row=7, column=column, value=value)
        cell.fill = PatternFill('solid', fgColor='FFF2CC')
        cell.alignment = Alignment(vertical='top', wrap_text=True)

    thin_gray = Side(style='thin', color='D9E2F3')
    for row in worksheet.iter_rows(min_row=6, max_row=7, min_col=1, max_col=len(KIB_HEADERS)):
        for cell in row:
            cell.border = Border(bottom=thin_gray)

    widths = [22, 28, 22, 30, 24, 20, 10, 12, 38, 14, 14, 18, 18, 24, 18, 38]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[_column_name(column)].width = width

    worksheet.row_dimensions[6].height = 38
    worksheet.row_dimensions[7].height = 42
    worksheet.freeze_panes = 'A7'
    worksheet.auto_filter.ref = f'A6:{last_column}7'

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _column_name(number):
    result = ''
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result
