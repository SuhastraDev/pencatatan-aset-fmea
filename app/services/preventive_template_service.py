"""Generate an Excel template for the preventive maintenance import."""

from datetime import date
from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HEADERS = ['No.', 'Tanggal', 'Nama Alat', 'Merk', 'SN', 'Type', 'Hasil', 'Ket.']
MONTHS_ID = (
    'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
    'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember',
)


def generate_preventive_template(rooms, template_date=None):
    """Return a workbook with accessible room sheets and one safe example row."""
    rooms = sorted(list(rooms), key=lambda room: (room.room_name or '').lower())
    template_date = template_date or date.today()

    workbook = Workbook()
    workbook.remove(workbook.active)

    if not rooms:
        rooms = [type('TemplateRoom', (), {'room_name': 'TEMPLATE'})()]

    used_titles = set()
    for room_index, room in enumerate(rooms):
        room_name = room.room_name or 'TEMPLATE'
        worksheet = workbook.create_sheet(_sheet_title(room_name, used_titles))
        _format_sheet(worksheet, room_name, template_date, include_example=room_index == 0)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _format_sheet(worksheet, room_name, template_date, include_example):
    worksheet.merge_cells('A1:H1')
    worksheet['A1'] = 'TEMPLATE PREVENTIVE MAINTENANCE'
    worksheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    worksheet['A1'].fill = PatternFill('solid', fgColor='0B7896')
    worksheet['A1'].alignment = Alignment(horizontal='center')

    worksheet.merge_cells('A2:H2')
    worksheet['A2'] = 'Rumah Sakit Khusus Gigi dan Mulut Provinsi Sumatera Selatan'
    worksheet['A2'].font = Font(italic=True, color='666666')
    worksheet['A2'].alignment = Alignment(horizontal='center')

    worksheet['A4'] = f'Ruangan : {room_name}'
    worksheet['A5'] = f'Tanggal : {_format_date(template_date)}'
    worksheet['A4'].font = worksheet['A5'].font = Font(bold=True, color='404040')

    for column, header in enumerate(HEADERS, start=1):
        cell = worksheet.cell(row=7, column=column, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E79')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    if include_example:
        example = [
            1,
            template_date,
            'CONTOH - Dental Unit',
            'Contoh Merk',
            'CONTOH-SN-001',
            'Contoh Type',
            'Baik',
            'Hapus baris CONTOH ini, lalu isi data asli.',
        ]
        for column, value in enumerate(example, start=1):
            cell = worksheet.cell(row=9, column=column, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.fill = PatternFill('solid', fgColor='FFF2CC')
    thin_gray = Side(style='thin', color='D9E2F3')
    for row in worksheet.iter_rows(min_row=7, max_row=9, min_col=1, max_col=8):
        for cell in row:
            cell.border = Border(bottom=thin_gray)

    widths = [8, 14, 30, 22, 20, 20, 35, 42]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + column)].width = width
    worksheet.row_dimensions[7].height = 24
    worksheet.row_dimensions[9].height = 38
    worksheet.freeze_panes = 'A8'
    worksheet.auto_filter.ref = 'A7:H9'


def _sheet_title(room_name, used_titles):
    base = re.sub(r'[:\\/?*\[\]]', '-', str(room_name)).strip() or 'TEMPLATE'
    base = base[:31]
    title = base
    suffix = 2
    while title in used_titles:
        suffix_text = f' ({suffix})'
        title = f'{base[:31 - len(suffix_text)]}{suffix_text}'
        suffix += 1
    used_titles.add(title)
    return title


def _format_date(value):
    return f'{value.day} {MONTHS_ID[value.month - 1]} {value.year}'
